import openpyxl
import datetime
import calendar
from notifier import Notifier
import os

def parse_borrowings(file_path):
    if not os.path.exists(file_path):
        return []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        borrowings = []
        last_bank = None
        last_product_type = None
        
        for row in sheet.iter_rows(min_row=6, values_only=True):
            # Update last known values if current row has them
            if row[2] and str(row[2]).strip() not in ['', '계']:
                last_bank = row[2]
            if row[3]:
                last_product_type = row[3]
                
            # Skip total rows or empty rows
            if row[2] == '계' or all(x is None for x in row):
                continue
                
            # Valid row must have an account number (row[4]) and a numeric amount (row[5])
            if row[4] and last_bank:
                amount = row[5]
                # Check if amount is numeric (to skip rows that are notes or shifted dates)
                if not isinstance(amount, (int, float)):
                    continue
                    
                try:
                    item = {
                        'bank': last_bank,
                        'product_type': last_product_type or row[3],
                        'amount': amount,
                        'rate': row[13],
                        'maturity_date': row[16],
                        'execution_date': row[15],
                        'account_no': str(row[4]),
                        'currency': 'KRW'
                    }
                    borrowings.append(item)
                except: continue
        return borrowings
    except: return []

def parse_financial_products(file_path):
    if not os.path.exists(file_path):
        return []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        products = []
        last_bank = None
        
        # Header is at row 1
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                last_bank = row[0]
            
            # Skip empty rows or rows without a product name
            if not row[1] or all(x is None for x in row):
                continue
                
            try:
                item = {
                    'bank': last_bank or row[0],
                    'product_name': row[1],
                    'type': row[2], # 수시/기간
                    'currency': row[3],
                    'amount': row[4],
                    'rate': row[5],
                    'maturity_date': row[7]
                }
                products.append(item)
            except: continue
        return products
    except: return []

def check_deadlines():
    notifier = Notifier()
    today = datetime.datetime.now().date()
    target_d_days = [30, 10, 3, 0]
    
    # 1. Check Borrowings
    borrowings = parse_borrowings("borrowings.xlsx")
    for b in borrowings:
        # Maturity
        if b['maturity_date'] and isinstance(b['maturity_date'], (datetime.datetime, datetime.date)):
            maturity = b['maturity_date']
            if isinstance(maturity, datetime.datetime): maturity = maturity.date()
            diff = (maturity - today).days
            if diff in target_d_days:
                notifier.notify_dday(f"차입금 만기 ({b['product_type']})", maturity.strftime("%Y-%m-%d"), f"D-{diff}" if diff > 0 else "D-Day", b['amount'] or 0, b['bank'])
        
        # Interest
        if b['execution_date'] and isinstance(b['execution_date'], (datetime.datetime, datetime.date)):
            exec_date = b['execution_date']
            if isinstance(exec_date, datetime.datetime): exec_date = exec_date.date()
            exec_day = exec_date.day
            next_m, next_y = (today.month + 1, today.year) if today.day > exec_day else (today.month, today.year)
            if next_m > 12: next_m, next_y = 1, next_y + 1
            _, last_day = calendar.monthrange(next_y, next_m)
            next_interest_date = datetime.date(next_y, next_m, min(exec_day, last_day))
            diff = (next_interest_date - today).days
            if diff in target_d_days:
                notifier.notify_dday("차입금 이자 지급일", next_interest_date.strftime("%Y-%m-%d"), f"D-{diff}" if diff > 0 else "D-Day", 0, b['bank'])

    # 2. Check Financial Products
    products = parse_financial_products("financial_products.xlsx")
    for p in products:
        if p['maturity_date'] and isinstance(p['maturity_date'], (datetime.datetime, datetime.date, str)):
            if p['maturity_date'] == "-": continue
            maturity = p['maturity_date']
            if isinstance(maturity, str):
                try: maturity = datetime.datetime.strptime(maturity, "%Y-%m-%d").date()
                except: continue
            elif isinstance(maturity, datetime.datetime): maturity = maturity.date()
            
            diff = (maturity - today).days
            if diff in target_d_days:
                amount_str = f"{p['amount']:,.0f} {p['currency']}"
                notifier.notify_dday(f"금융상품 만기 ({p['product_name']})", maturity.strftime("%Y-%m-%d"), f"D-{diff}" if diff > 0 else "D-Day", p['amount'], f"{p['bank']} ({p['currency']})")

if __name__ == "__main__":
    print(f"--- Running Treasury Auto Check ({datetime.datetime.now()}) ---")
    check_deadlines()
    print("--- Check Completed ---")
